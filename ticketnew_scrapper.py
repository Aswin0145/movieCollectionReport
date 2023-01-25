
from openpyxl import Workbook
import requests
import json
import locale
import time
from datetime import datetime
locale.setlocale(locale.LC_MONETARY, 'en_IN')



# The below code is to fetch cities available in paytm (Ticket new)

# s =requests.session()
# headers = {"user-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36"}
# r = s.get("https://paytm.com/movies/varisu-movie-detail-154763?frmtid=9litjdoar", headers = headers)
# res = r.text
# obj_cities=json.loads('{"cities":'+res.split(',"topCities"')[0].split('{"cities":')[1]+'}')['cities']

# cities =[]
# for obj in obj_cities:
#     cities.append(obj['value'])


# If u want only cities in Tamilnadu uncomment below code

# cities_in_tn = pd.read_csv('Cities and Towns in Tamil Nadu - Population statistics.csv')
# cities_in_tn = cities_in_tn['Name'].to_list()
# filtered_city = []
# for city in cities_in_tn:
#     if city.lower() in cities:
#         filtered_city.append(city.lower())

# If u want all over india uncomment below code

# filtered_city = [city.lower() for city in cities]
        
# print(filtered_city)
 
filtered_city = ['acharapakkam', 'alanganallur', 'alangayam', 'alangudi', 'ambasamudram', 'ambur', 'ammapettai', 'ammapettai', 'annur', 'anthiyur', 'aranthangi', 'arcot', 'ariyalur', 'arumbavur', 'aruppukkottai', 'attur', 'batlagundu', 'belur', 'bhuvanagiri', 'chidambaram', 'chinnamanur', 'chennai', 'chinnasalem', 'coimbatore', 'cuddalore', 'devakottai', 'dharapuram', 'dharmapuri', 'dindigul', 'erode', 'gingee', 'gobichettipalayam', 'gudiyatham', 'harur', 'hosur', 'jalakandapuram', 'kadayam', 'kambainallur', 'kangeyam', 'karur', 'katpadi', 'kaveripattinam', 'keeranur', 'keeranur', 'kinathukadavu', 'kolathur', 'kondur', 'kovilpatti', 'krishnagiri', 'kulithalai', 'kumbakonam', 'kurinjipadi', 'kuzhithurai', 'lalgudi', 'madurai', 'maduranthakam', 'manapparai', 'mannargudi', 'mayiladuthurai', 'mettur', 'musiri', 'muthur', 'nagapattinam', 'nagercoil', 'namakkal', 'nambiyur', 'neyveli', 'palani', 'pallipalayam', 'panruti', 'pattukkottai', 'pennagaram', 'peravurani', 'periyakulam', 'perundurai', 'pollachi', 'ponnamaravathi', 'ponneri', 'pudukkottai', 'pudukkottai', 'rajapalayam', 'rajapalayam', 'ramanathapuram', 'rasipuram', 'salem', 'sathankulam', 'sathyamangalam', 'sattur', 'sivakasi', 'srivaikuntam', 'srivilliputhur', 'surandai', 'tenkasi', 'thammampatti', 'thanjavur', 'thiruthuraipoondi', 'thiruvarur', 'thuraiyur', 'tindivanam', 'tiruchengode', 'tirunelveli', 'tirupattur', 'uthangarai', 'vadalur', 'vadipatti', 'vaniyambadi', 'vedaranyam', 'vellakoil', 'vellore', 'venkatapuram', 'vikravandi', 'virudhachalam']
movie=input("Thunivu or Varisu: ")

if movie=="Varisu":
    movie_code="9litjdoar"
elif movie=="Thunivu":
    movie_code="o_7tzoqjs"
else:
    movie_code = input("Enter The movie code: ")
    
today = datetime.now()
    
movie_date = input("Enter the date (yyyy-mm-dd). Enter only current date or upcoming date: ")

user_date = datetime.strptime(movie_date,"%Y-%m-%d")

if user_date < today:
    movie_date=today.strftime("%Y-%m-%d")
    print(f"Since tou enter date below today. We have taken today's date ({movie_date})")
 
    
column_headers = ["City", "Total Theatres","Total Shows","Total Seats","Total Booked Seats","Total Occupancy","Total amt collected"]
workbook = Workbook()
sheet = workbook.active
     
total_shows = 0
total_seats = 0
total_seats_booked= 0
total_amt_collected_full= 0
total_theatres =0 

row = 1

for col,column_header in enumerate(column_headers,start=1):
    sheet.cell(row,col).value = column_header

for city in filtered_city:
    s =requests.session()
    headers = {"user-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36"}
    r = s.get(f"https://apiproxy.paytm.com/v3/movies/search/movie?meta=1&reqData=1&city={city}&movieCode={movie_code}&date={movie_date}&version=3&site_id=1&channel=web&child_site_id=1", headers = headers)
    
    print(city+"(status code) : "+str(r.status_code))
    if r.status_code == 200 :
        res = json.loads(r.text)
        cinemas = res["meta"]["cinemas"]
        sessions = res["pageData"]["sessions"]
        total_no_seats_booked_in_day = 0
        total_no_of_seats_in_a_day = 0
        total_amt_collected = 0
        total_shows_in_a_day = 0
        for cinema in cinemas:
            try:
                cinema_session= sessions[str(cinema["id"])]
                total_seats_booked_for_the_day = 0
                total_seats_in_shows_for_the_day = 0
                total_amt_of_booked_seat =0
                total_show_in_a_cinema = 0
            except:
                print("May be key Error")

            for show in cinema_session:
                total_seats_in_shows_for_the_day = total_seats_in_shows_for_the_day + show['total']
                total_seats_booked_for_the_day = total_seats_booked_for_the_day + show['total']- show['avail']
                
                seating_areas = show['areas']
                for seat_class in seating_areas:
                    total_amt_of_booked_seat = total_amt_of_booked_seat + (seat_class['sTotal']-seat_class['sAvail'])*seat_class['price']
                    total_show_in_a_cinema = total_show_in_a_cinema + 1
            
            total_no_seats_booked_in_day = total_no_seats_booked_in_day+total_seats_booked_for_the_day
            total_no_of_seats_in_a_day= total_no_of_seats_in_a_day + total_seats_in_shows_for_the_day
            total_amt_collected = total_amt_collected + total_amt_of_booked_seat
            total_shows_in_a_day = total_shows_in_a_day + total_show_in_a_cinema
        
        print("city: "+city)
        print("Total no of theatres in a day: "+str(len(cinemas)))
        print("Total no of shows in a day: "+str(total_shows_in_a_day))    
        print("Total no of seats in a day: " +str(total_no_of_seats_in_a_day))
        print("Total no of booked seats in a day: "+str(total_no_seats_booked_in_day))
        print("percentage occupancy of the day: "+str(round((total_no_seats_booked_in_day/total_no_of_seats_in_a_day)*100,2))+"%")
        print("Total amt collected based in a day: "+ str(locale.currency(round(total_amt_collected,2),grouping=True)))
        
        row = row + 1
        sheet.cell(row,1).value = city
        sheet.cell(row,2).value = len(cinemas)
        sheet.cell(row,3).value = total_shows_in_a_day
        sheet.cell(row,4).value = total_no_of_seats_in_a_day
        sheet.cell(row,5).value = total_no_seats_booked_in_day
        sheet.cell(row,6).value = str(round((total_no_seats_booked_in_day/total_no_of_seats_in_a_day)*100,2))+"%"
        sheet.cell(row,7).value = str(locale.currency(round(total_amt_collected,2),grouping=True))
        
        
        total_theatres = total_theatres + len(cinemas)
        total_shows = total_shows + total_shows_in_a_day
        total_seats = total_seats + total_no_of_seats_in_a_day
        total_seats_booked = total_seats_booked + total_no_seats_booked_in_day
        total_amt_collected_full = total_amt_collected_full + total_amt_collected
        time.sleep(10)

row = row + 2
sheet.cell(row,1).value = "Total"
sheet.cell(row,2).value = total_theatres
sheet.cell(row,3).value = total_shows
sheet.cell(row,4).value = total_seats
sheet.cell(row,5).value = total_seats_booked
sheet.cell(row,6).value = str(round((total_seats_booked/total_seats)*100,2))+"%"
sheet.cell(row,7).value = str(locale.currency(round(total_amt_collected_full,2),grouping=True))  
workbook.save(movie+str(movie_date)+".xlsx")      
print("Total no of theatres in a day: "+str(total_theatres))
print("Total no of shows in a day: "+str(total_shows))    
print("Total no of seats in a day: " +str(total_seats))
print("Total no of booked seats in a day: "+str(total_seats_booked))
print("percentage occupancy of the day: "+str(round((total_seats_booked/total_seats)*100,2))+"%")
print("Total amt collected based in a day: "+ str(locale.currency(round(total_amt_collected_full,2),grouping=True)))    